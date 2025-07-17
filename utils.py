# utils.py

from openpyxl.utils import get_column_letter  # 列番号 (1, 2, 3...) を列名 (A, B, C...) に変換する機能
from openpyxl.worksheet.worksheet import Worksheet # 型ヒント（関数の引数がどの型かを明示）のために使用
import pandas as pd # データ分析ライブラリ。主にデータ構造として使用

def apply_summary_sheet_styles(ws: Worksheet, df: pd.DataFrame):
    """
    データフレームの内容に基づいて、Excelシートの列幅を自動調整する。
    日本語のような全角文字は幅を広く計算し、見やすいレイアウトを作成する。

    Args:
        ws (Worksheet): スタイルを適用する対象のopenpyxlワークシートオブジェクト。
        df (pd.DataFrame): 列幅計算の元となるデータを含むpandasデータフレーム。
    """
    # enumerateを使い、列のインデックス番号(i)と列名(column_name)を同時に取得してループ
    # openpyxlの列番号は1から始まるため、start=1 を指定
    for i, column_name in enumerate(df.columns, 1):
        
        # --- 1. ヘッダーの長さを基準に、列幅の初期値を計算 ---
        try:
            # 日本語（全角文字）は半角文字より幅が広いため、3文字分として計算する。
            # '一'から'龠'はJIS第一・第二水準漢字、'ぁ'から'ん'はひらがな、'Ａ'から'ｚ'は全角英数字をカバー。
            jp_char_count = sum(1 for char in str(column_name) if '一' <= char <= '龠' or 'ぁ' <= char <= 'ん' or 'ァ' <= char <= 'ン' or 'Ａ' <= char <= 'ｚ')
            # (全体の文字数 - 全角文字数) + (全角文字数 * 3) で、おおよその表示幅を計算。
            # +2 は余白（パディング）分。
            max_length = (len(str(column_name)) - jp_char_count) + (jp_char_count * 3) + 2
        except:
            # ヘッダーが予期せぬデータ型で上記の計算が失敗した場合のエラー処理。
            # プログラムが停止しないよう、単純な文字数×1.5を代替値とする。
            max_length = len(str(column_name)) * 1.5

        # --- 2. 列内の各セルの長さを確認し、最大幅を更新 ---
        # データフレームの当該列から、欠損値(NaN)を除いた各セルのデータをループ処理。
        for cell_data in df[column_name].dropna():
            try:
                # ヘッダーと同様に、セル内の全角文字を考慮して表示幅を計算。
                jp_char_count_cell = sum(1 for char in str(cell_data) if '一' <= char <= '龠' or 'ぁ' <= char <= 'ん' or 'ァ' <= char <= 'ン' or 'Ａ' <= char <= 'ｚ')
                cell_len = (len(str(cell_data)) - jp_char_count_cell) + (jp_char_count_cell * 3) + 2
                
                # これまでの最大幅(max_length)より現在のセルの幅が広ければ、最大幅を更新。
                if cell_len > max_length:
                    max_length = cell_len
            except:
                # セル内に計算不能なデータ型（例: 特殊なオブジェクト）が含まれていた場合のエラー処理。
                # スタイル適用はベストエフォートとし、問題のセルはスキップして処理を続行する。
                pass
                
        # --- 3. 計算した最大幅を列に適用 ---
        # 列幅が50を超えると見づらくなるため、上限を50に設定。
        adjusted_width = min(max_length, 500)
        # 列番号(i)を 'A', 'B', 'C'... といったExcelの列名に変換。
        column_letter = get_column_letter(i)
        # ワークシートオブジェクト(ws)の列の寸法(column_dimensions)に、計算した幅を設定。
        ws.column_dimensions[column_letter].width = adjusted_width






























# # utils.py

# from openpyxl.utils import get_column_letter
# from openpyxl.worksheet.worksheet import Worksheet
# import pandas as pd

# def apply_summary_sheet_styles(ws: Worksheet, df: pd.DataFrame):
#     """
#     データフレームの内容に基づいてExcelシートの列幅を自動調整する。
    
#     Args:
#         ws (Worksheet): 対象のOpenpyxlワークシートオブジェクト。
#         df (pd.DataFrame): スタイル適用の元となるデータフレーム。
#     """
#     for i, column_name in enumerate(df.columns, 1):
#         # ヘッダーの長さを初期値とする
#         # 日本語は2文字としてカウント
#         try:
#             header_len_jp = sum(1 for char in str(column_name) if '一' <= char <= '龠' or 'ぁ' <= char <= 'ん' or 'Ａ' <= char <= 'ｚ')
#             max_length = (len(str(column_name)) - header_len_jp) + (header_len_jp * 2) + 2
#         except:
#             max_length = len(str(column_name)) * 1.5

#         # 各セルのデータ長を比較して最大値を求める
#         for cell_data in df[column_name].dropna():
#             try:
#                 # 日本語文字は2バイトとして計算
#                 len_jp = sum(1 for char in str(cell_data) if '一' <= char <= '龠' or 'ぁ' <= char <= 'ん' or 'Ａ' <= char <= 'ｚ')
#                 cell_len = (len(str(cell_data)) - len_jp) + (len_jp * 2) + 2
#                 if cell_len > max_length:
#                     max_length = cell_len
#             except:
#                 # 計算できないデータはスキップ
#                 pass
                
#         # 列幅を設定（上限を50とする）
#         adjusted_width = min(max_length, 50)
#         column_letter = get_column_letter(i)
#         ws.column_dimensions[column_letter].width = adjusted_width
