



# repositories/excel_repository.py
"""
Excelファイルの永続化（読み書き）に特化したリポジトリモジュール。
openpyxlライブラリの具体的な操作をこのクラス内にカプセル化（閉じ込める）する。
"""

import os
import errno  # ディスク空き容量不足など、OSレベルのエラーコードを判定するために使用
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.exceptions import InvalidFileException  # Excelファイル破損などを検知
import pandas as pd
from typing import List

# アプリケーション内の他モジュールをインポート
import config
from utils import apply_summary_sheet_styles
from models.student import Student

class ExcelRepository:
    """
    Excelファイルへの全ての読み書きアクセスを管理するクラス。
    """
    def __init__(self, file_path: str):
        """
        コンストラクタ。指定されたパスのExcelファイルを読み込んで準備する。
        
        【エラーハンドリング】
        ファイルが存在しない、破損している、非対応形式であるといった、
        ファイルを開く段階での問題をここで検知し、具体的なエラーを発生させる。
        """
        self.file_path: str = file_path
        try:
            # openpyxlでExcelファイルを開き、workbookオブジェクトとして保持する
            self.workbook = load_workbook(file_path)
        except FileNotFoundError:
            # 指定されたパスにファイルが存在しない場合
            raise FileNotFoundError(f"指定されたExcelファイルが見つかりません: {file_path}")
        except InvalidFileException:
            # ファイルが破損している、または古い.xls形式など、openpyxlが対応していない形式の場合
            raise IOError(f"Excelファイルを開けません。\nファイルが破損しているか、サポートされていない形式（例: .xls）の可能性があります。\nファイル: {os.path.basename(file_path)}")
        except Exception as e:
             # 上記以外の予期せぬ読み込みエラー
             raise IOError(f"Excelファイルの読み込み中に予期せぬエラーが発生しました。\nファイル: {os.path.basename(file_path)}\n詳細: {e}")

    def save(self):
        """
        ここまでの変更をExcelファイルに上書き保存する。
        
        【エラーハンドリング】
        ファイルへの書き込み権限がない（他のアプリで開かれている）、ディスクの空き容量がない
        といった、保存時の問題をここで検知する。
        """
        try:
            self.workbook.save(self.file_path)
        except PermissionError:
            # ファイルが他のプログラム（例: Excel本体）で開かれていてロックされている場合
            raise PermissionError(f"Excelファイルへの保存に失敗しました。\nファイルが他のプログラムで開かれていないか、書き込み権限があるか確認してください。\nファイル: {self.file_path}")
        except OSError as e:
            # ディスクI/Oに関するOSレベルのエラー
            if e.errno == errno.ENOSPC:  # ENOSPCは「No space left on device」を示すエラーコード
                raise IOError(f"ディスクの空き容量が不足しているため、ファイルを保存できません。\nファイル: {self.file_path}")
            else:
                raise IOError(f"ファイルの保存中にOSエラーが発生しました。\n詳細: {e}")
        except Exception as e:
            # その他の予期せぬ保存エラー
            raise IOError(f"Excelファイルへの保存中に予期せぬエラーが発生しました。\n詳細: {e}")

    def update_grades_sheet(self, term: str, students: List[Student], status_callback):
        """
        「成績一覧」シートをCSVデータに基づいて更新する。
        """
        # --- 1. 対象シートの特定と事前チェック ---
        sheet_name = config.SHEET_NAME_GRADES_TEMPLATE.format(term=term)
        # シートが存在しない場合は処理をスキップ
        if sheet_name not in self.workbook.sheetnames:
            status_callback(f"警告: シート '{sheet_name}' が見つかりません。スキップします。")
            return

        ws = self.workbook[sheet_name]

        # ヘッダーが4行構成であることを前提としているため、行数が不足している場合は処理をスキップ
        if ws.max_row < 4:
            status_callback(f"警告: シート '{sheet_name}' のヘッダー情報が不足しています（4行未満）。処理をスキップします。")
            return

        status_callback(f"処理中: シート '{sheet_name}'")

        # --- 2. 書き込み位置を特定するための準備 ---
        # Excelシートの複雑なヘッダーを解析し、どの科目が何列目にあるかの対応表（辞書）を作成する
        subject_column_map = self._map_subject_columns(ws)
        other_column_map = self._map_other_columns(ws)

        # 既存の学生IDとExcel上の行番号の対応表を作成し、更新処理を高速化する
        student_row_map = {str(ws.cell(row=r, column=1).value): r for r in range(5, ws.max_row + 1) if ws.cell(row=r, column=1).value}
        # 新規学生を追記する場合の開始行を計算
        next_new_student_row = (max(student_row_map.values()) + 1) if student_row_map else 5

        # --- 3. 学生データに基づき、セルを更新または追記 ---
        for student in students:
            # 既存学生か新規学生かを判定
            target_row = student_row_map.get(student.id)
            if not target_row:
                target_row = next_new_student_row
                next_new_student_row += 1

            # セルに値を書き込み
            ws.cell(row=target_row, column=1, value=student.id)
            ws.cell(row=target_row, column=2, value=student.name)

            for (subject, test_type), score in student.scores.items():
                col = subject_column_map.get((subject, test_type))
                if col: ws.cell(row=target_row, column=col, value=score)

            for key, value in student.summary_data.items():
                col = other_column_map.get(key)
                if col and key != "再試数": ws.cell(row=target_row, column=col, value=value)

            # 「再試数」列に再試の数をカウントする数式(=COUNT)を自動入力
            retest_count_col = other_column_map.get("再試数")
            retest_cols = [c for (s, t), c in subject_column_map.items() if t == config.KEY_TEST_TYPES[1]]
            if retest_count_col and retest_cols:
                cell_addresses = [f"{get_column_letter(c)}{target_row}" for c in retest_cols]
                ws.cell(row=target_row, column=retest_count_col, value=f"=COUNT({','.join(cell_addresses)})")

        status_callback(f"-> シート '{sheet_name}' の更新完了。")


    def create_summary_sheet(self, sheet_name: str, headers: List[str], data_rows: List[list]):
        """
        新しい集計シートを作成し、ヘッダーとデータ行を一括で書き込む。
        """
        # もし同名のシートが既に存在すれば、一度削除して新しいものを作成する
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
        ws = self.workbook.create_sheet(sheet_name)

        ws.append(headers) # ヘッダー行を書き込み
        self.style_row(sheet_name, 1, is_header=True) # ヘッダースタイルを適用

        for row in data_rows:
            ws.append(row) # データ行を1行ずつ書き込み

        # 列幅を自動調整
        if data_rows:
            df_for_styling = pd.DataFrame(data_rows, columns=headers)
            apply_summary_sheet_styles(ws, df_for_styling)

    # --- 以下、シートの見た目を整えるための補助的なメソッド群 ---

    def style_row(self, sheet_name: str, row_number: int, is_header: bool = False):
        """指定した行にスタイル（背景色など）を適用する"""
        if sheet_name not in self.workbook.sheetnames: return
        ws = self.workbook[sheet_name]
        if is_header:
            font = Font(bold=True, color=config.HEADER_FONT_COLOR)
            fill = PatternFill(start_color=config.HEADER_FILL_COLOR, end_color=config.HEADER_FILL_COLOR, fill_type="solid")
            for cell in ws[row_number]: cell.font = font; cell.fill = fill

    def merge_header_cells(self, sheet_name: str, start_row: int, start_col: int, end_col: int):
        """指定された範囲のセルを結合する"""
        if sheet_name not in self.workbook.sheetnames: return
        self.workbook[sheet_name].merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)

    def align_header_center(self, sheet_name: str, row_start: int, row_end: int):
        """指定された行範囲のセルを中央揃えにする"""
        if sheet_name not in self.workbook.sheetnames: return
        ws = self.workbook[sheet_name]
        center_alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=row_start, max_row=row_end):
            for cell in row: cell.alignment = center_alignment

    def apply_borders_to_all_cells(self, sheet_name: str):
        """指定されたシートのデータが存在する全てのセルに罫線を適用する"""
        if sheet_name not in self.workbook.sheetnames: return
        ws = self.workbook[sheet_name]
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows():
            for cell in row: cell.border = thin_border

    # --- 以下、内部でのみ使用されるプライベートメソッド群 ---

    def _map_subject_columns(self, ws: Worksheet) -> dict:
        """成績一覧シートのヘッダーを解析し、(科目名, 種別) -> 列番号 の辞書を作成する"""
        mapping = {}
        for c_idx, cell in enumerate(ws[4], 1): # ヘッダー4行目（本試/再試など）を走査
            if cell.value in config.KEY_TEST_TYPES:
                # 結合されたセルを考慮して、1行目から科目名を取得
                subject_name = self._get_merged_cell_value(ws, 1, c_idx)
                if subject_name: mapping[(subject_name, cell.value)] = c_idx
        return mapping

    def _map_other_columns(self, ws: Worksheet) -> dict:
        """成績一覧シートのヘッダーを解析し、項目名 -> 列番号 の辞書を作成する"""
        mapping = {}
        for c_idx, cell in enumerate(ws[1], 1): # ヘッダー1行目（総点など）を走査
             if cell.value in config.KEY_OTHER_COLS: mapping[cell.value] = c_idx
        return mapping

    def _get_merged_cell_value(self, ws: Worksheet, row: int, col: int) -> str:
        """
        指定されたセルが結合されている場合を考慮して、そのセルの表示上の値を取得する。
        セルが結合されていると、左上のセル以外は値がNoneになるため、
        Noneの場合は左のセルを順に遡って値を探す。
        """
        cell = ws.cell(row=row, column=col)
        if cell.value is not None: return cell.value
        # 値がNoneの場合、左のセルを順に探索
        for c_idx in range(col - 1, 0, -1):
            val = ws.cell(row=row, column=c_idx).value
            if val is not None: return val
        return None































# import os
# import errno
# from openpyxl import load_workbook
# from openpyxl.worksheet.worksheet import Worksheet
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils.exceptions import InvalidFileException
# import pandas as pd
# from typing import List

# # アプリケーション内の他モジュールをインポート
# import config
# from utils import apply_summary_sheet_styles
# from models.student import Student


# class ExcelRepository:
#     """
#     Excelファイルの読み書き（永続化）に特化したクラス。
#     openpyxlライブラリの操作をこのクラス内に閉じ込める。
#     """
#     def __init__(self, file_path: str):
#         self.file_path: str = file_path
#         try:
#             # openpyxlでExcelファイルを開く
#             self.workbook = load_workbook(file_path)
#         except FileNotFoundError:
#             # ファイルが存在しない場合は専用のエラーを投げる
#             raise FileNotFoundError(f"指定されたExcelファイルが見つかりません: {file_path}")
#         except InvalidFileException:
#             # ファイルが破損している、または非対応形式の場合
#             raise IOError(f"Excelファイルを開けません。\nファイルが破損しているか、サポートされていない形式（例: .xls）の可能性があります。\nファイル: {os.path.basename(file_path)}")
#         except Exception as e:
#              # その他の予期せぬ読み込みエラー
#              raise IOError(f"Excelファイルの読み込み中に予期せぬエラーが発生しました。\nファイル: {os.path.basename(file_path)}\n詳細: {e}")

#     def save(self):
#         """ここまでの変更をExcelファイルに上書き保存する。"""
#         try:
#             self.workbook.save(self.file_path)
#         except PermissionError:
#             raise PermissionError(f"Excelファイルへの保存に失敗しました。\nファイルが他のプログラムで開かれていないか、書き込み権限があるか確認してください。\nファイル: {self.file_path}")
#         except OSError as e:
#             if e.errno == errno.ENOSPC: # ENOSPCは「No space left on device」を示す
#                 raise IOError(f"ディスクの空き容量が不足しているため、ファイルを保存できません。\nファイル: {self.file_path}")
#             else:
#                 raise IOError(f"ファイルの保存中にOSエラーが発生しました。\n詳細: {e}")
#         except Exception as e:
#             raise IOError(f"Excelファイルへの保存中に予期せぬエラーが発生しました。\n詳細: {e}")

#     def update_grades_sheet(self, term: str, students: List[Student], status_callback):
#         """
#         「成績一覧」シートをCSVデータに基づいて更新する。
#         再試数をカウントする数式も自動入力する。
#         """
#         # configからテンプレートを読み込み、対象のシート名を生成
#         sheet_name = config.SHEET_NAME_GRADES_TEMPLATE.format(term=term)
#         # シートが存在しない場合は警告を出して処理を中断
#         if sheet_name not in self.workbook.sheetnames:
#             status_callback(f"警告: シート '{sheet_name}' が見つかりません。スキップします。")
#             return

#         ws = self.workbook[sheet_name]

#         # ヘッダー行が存在するか簡易チェック
#         if ws.max_row < 4:
#             status_callback(f"警告: シート '{sheet_name}' のヘッダー情報が不足しています（4行未満）。処理をスキップします。")
#             return

#         status_callback(f"処理中: シート '{sheet_name}'")

#         # Excelシートのヘッダーを解析し、書き込むべき列を特定するマッピングを作成
#         subject_column_map = self._map_subject_columns(ws)
#         other_column_map = self._map_other_columns(ws)

#         retest_cols = [
#             col_num for (subject, test_type), col_num in subject_column_map.items()
#             if test_type == config.KEY_TEST_TYPES[1]  # KEY_TEST_TYPESの '再試'
#         ]
#         retest_count_col = other_column_map.get("再試数")
#         student_row_map = {}
#         for r_idx in range(5, ws.max_row + 1):
#             cell_val = ws.cell(row=r_idx, column=1).value
#             if cell_val is not None and str(cell_val).strip():
#                 student_row_map[str(cell_val)] = r_idx
#         next_new_student_row = (max(student_row_map.values()) + 1) if student_row_map else 5
#         for student in students:
#             target_row = student_row_map.get(student.id)
#             if not target_row:
#                 target_row = next_new_student_row
#                 next_new_student_row += 1
#             ws.cell(row=target_row, column=1, value=student.id)
#             ws.cell(row=target_row, column=2, value=student.name)
#             for (subject, test_type), score in student.scores.items():
#                 col = subject_column_map.get((subject, test_type))
#                 if col:
#                     ws.cell(row=target_row, column=col, value=score)
#             for key, value in student.summary_data.items():
#                 if key == "再試数":
#                     continue
#                 col = other_column_map.get(key)
#                 if col:
#                     ws.cell(row=target_row, column=col, value=value)
#             if retest_count_col and retest_cols:
#                 cell_addresses = [f"{get_column_letter(col)}{target_row}" for col in retest_cols]
#                 formula = f"=COUNT({','.join(cell_addresses)})"
#                 ws.cell(row=target_row, column=retest_count_col, value=formula)
#         status_callback(f"-> シート '{sheet_name}' の更新完了。")


#     def create_summary_sheet(self, sheet_name: str, headers: List[str], data_rows: List[list]):
#         """
#         新しい集計シートを作成し、ヘッダーとデータ行を一括で書き込む。
#         """
#         if sheet_name in self.workbook.sheetnames:
#             del self.workbook[sheet_name]
#         ws = self.workbook.create_sheet(sheet_name)

#         ws.append(headers)
#         self.style_row(sheet_name, 1, is_header=True)

#         for row in data_rows:
#             ws.append(row)

#         if data_rows:
#             df_for_styling = pd.DataFrame(data_rows, columns=headers)
#             apply_summary_sheet_styles(ws, df_for_styling)

#     def style_row(self, sheet_name: str, row_number: int, is_header: bool = False):
#         """指定した行にスタイルを適用するヘルパーメソッド"""
#         if sheet_name not in self.workbook.sheetnames:
#             return
#         ws = self.workbook[sheet_name]

#         if is_header:
#             font = Font(bold=True, color=config.HEADER_FONT_COLOR)
#             fill = PatternFill(start_color=config.HEADER_FILL_COLOR, end_color=config.HEADER_FILL_COLOR, fill_type="solid")
#             for cell in ws[row_number]:
#                 cell.font = font
#                 cell.fill = fill

#     def merge_header_cells(self, sheet_name: str, start_row: int, start_col: int, end_col: int):
#         """指定された範囲のセルを結合する"""
#         if sheet_name not in self.workbook.sheetnames:
#             return
#         ws = self.workbook[sheet_name]
#         ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)

#     def align_header_center(self, sheet_name: str, row_start: int, row_end: int):
#         """指定された行範囲のセルを中央揃えにする"""
#         if sheet_name not in self.workbook.sheetnames:
#             return
#         ws = self.workbook[sheet_name]
#         center_alignment = Alignment(horizontal='center', vertical='center')
#         for row in ws.iter_rows(min_row=row_start, max_row=row_end):
#             for cell in row:
#                 cell.alignment = center_alignment

#     def apply_borders_to_all_cells(self, sheet_name: str):
#         """
#         指定されたシートのデータが存在する全てのセルに罫線を適用する。
#         """
#         if sheet_name not in self.workbook.sheetnames:
#             return
#         ws = self.workbook[sheet_name]

#         thin_border_side = Side(style='thin', color='000000')
#         border = Border(left=thin_border_side,
#                         right=thin_border_side,
#                         top=thin_border_side,
#                         bottom=thin_border_side)

#         for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#             for cell in row:
#                 cell.border = border

#     def _map_subject_columns(self, ws: Worksheet) -> dict:
#         """成績一覧シートのヘッダー（科目）を解析し、(科目名, 種別) -> 列番号 の辞書を作成する"""
#         mapping = {}
#         fourth_row = ws[4]
#         for cell_idx, cell_in_4th_row in enumerate(fourth_row):
#             test_type = cell_in_4th_row.value
#             if test_type in config.KEY_TEST_TYPES:
#                 col_num_1based = cell_idx + 1
#                 subject_name = self._get_merged_cell_value(ws, 1, col_num_1based)
#                 if subject_name:
#                     mapping[(subject_name, test_type)] = col_num_1based
#         return mapping

#     def _map_other_columns(self, ws: Worksheet) -> dict:
#         """成績一覧シートのヘッダー（総点など）を解析し、項目名 -> 列番号 の辞書を作成する"""
#         mapping = {}
#         first_row = ws[1]
#         for cell_idx, cell_in_1st_row in enumerate(first_row):
#              if cell_in_1st_row.value in config.KEY_OTHER_COLS:
#                  mapping[cell_in_1st_row.value] = cell_idx + 1
#         return mapping

#     def _get_merged_cell_value(self, ws: Worksheet, row: int, col: int) -> str:
#         """
#         指定されたセルが結合されている場合を考慮して、そのセルの表示上の値を取得する。
#         """
#         cell = ws.cell(row=row, column=col)
#         if cell.value is not None:
#             return cell.value

#         current_search_col = col - 1
#         while current_search_col >= 1:
#             header_cell_to_check = ws.cell(row=row, column=current_search_col)
#             if header_cell_to_check.value is not None:
#                 return header_cell_to_check.value
#             current_search_col -= 1
#         return None














































# # from openpyxl import load_workbook
# # from openpyxl.worksheet.worksheet import Worksheet
# # from openpyxl.utils import get_column_letter
# # from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# # import pandas as pd
# # from typing import List

# # # アプリケーション内の他モジュールをインポート
# # import config
# # from utils import apply_summary_sheet_styles
# # from models.student import Student


# # class ExcelRepository:
# #     """
# #     Excelファイルの読み書き（永続化）に特化したクラス。
# #     openpyxlライブラリの操作をこのクラス内に閉じ込める。
# #     """
# #     def __init__(self, file_path: str):
# #         self.file_path: str = file_path
# #         try:
# #             # openpyxlでExcelファイルを開く
# #             self.workbook = load_workbook(file_path)
# #         except FileNotFoundError:
# #             # ファイルが存在しない場合は専用のエラーを投げる
# #             raise FileNotFoundError(f"指定されたExcelファイルが見つかりません: {file_path}")

# #     def save(self):
# #         """ここまでの変更をExcelファイルに上書き保存する。"""
# #         self.workbook.save(self.file_path)

# #     def update_grades_sheet(self, term: str, students: List[Student], status_callback):
# #         """
# #         「成績一覧」シートをCSVデータに基づいて更新する。
# #         再試数をカウントする数式も自動入力する。
# #         """
# #         # configからテンプレートを読み込み、対象のシート名を生成
# #         sheet_name = config.SHEET_NAME_GRADES_TEMPLATE.format(term=term)
# #         # シートが存在しない場合は警告を出して処理を中断
# #         if sheet_name not in self.workbook.sheetnames:
# #             status_callback(f"警告: シート '{sheet_name}' が見つかりません。スキップします。")
# #             return

# #         ws = self.workbook[sheet_name]
# #         status_callback(f"処理中: シート '{sheet_name}'")

# #         # Excelシートのヘッダーを解析し、書き込むべき列を特定するマッピングを作成
# #         subject_column_map = self._map_subject_columns(ws) # 例: {('数学', '本試'): 5}
# #         other_column_map = self._map_other_columns(ws)   # 例: {'総点': 15, '再試数': 30}

# #         # 「再試」科目の列番号のリストを事前に作成
# #         retest_cols = [
# #             col_num for (subject, test_type), col_num in subject_column_map.items()
# #             if test_type == config.KEY_TEST_TYPES[1]  # KEY_TEST_TYPESの '再試'
# #         ]

# #         # 「再試数」列の列番号を取得
# #         retest_count_col = other_column_map.get("再試数")

# #         # Excelシートの既存の学生IDと行番号をマッピングする
# #         student_row_map = {}
# #         for r_idx in range(5, ws.max_row + 1):
# #             cell_val = ws.cell(row=r_idx, column=1).value
# #             if cell_val is not None and str(cell_val).strip():
# #                 student_row_map[str(cell_val)] = r_idx

# #         # 新規学生を追記する場合の開始行を計算
# #         next_new_student_row = (max(student_row_map.values()) + 1) if student_row_map else 5

# #         # 全ての学生データについてループ処理
# #         for student in students:
# #             target_row = student_row_map.get(student.id)
# #             if not target_row:
# #                 target_row = next_new_student_row
# #                 next_new_student_row += 1

# #             ws.cell(row=target_row, column=1, value=student.id)
# #             ws.cell(row=target_row, column=2, value=student.name)

# #             for (subject, test_type), score in student.scores.items():
# #                 col = subject_column_map.get((subject, test_type))
# #                 if col:
# #                     ws.cell(row=target_row, column=col, value=score)

# #             for key, value in student.summary_data.items():
# #                 if key == "再試数":
# #                     continue
# #                 col = other_column_map.get(key)
# #                 if col:
# #                     ws.cell(row=target_row, column=col, value=value)

# #             if retest_count_col and retest_cols:
# #                 cell_addresses = [f"{get_column_letter(col)}{target_row}" for col in retest_cols]
# #                 formula = f"=COUNT({','.join(cell_addresses)})"
# #                 ws.cell(row=target_row, column=retest_count_col, value=formula)

# #         status_callback(f"-> シート '{sheet_name}' の更新完了。")


# #     def create_summary_sheet(self, sheet_name: str, headers: List[str], data_rows: List[list]):
# #         """
# #         新しい集計シートを作成し、ヘッダーとデータ行を一括で書き込む。
# #         """
# #         if sheet_name in self.workbook.sheetnames:
# #             del self.workbook[sheet_name]
# #         ws = self.workbook.create_sheet(sheet_name)

# #         ws.append(headers)
# #         self.style_row(sheet_name, 1, is_header=True)

# #         for row in data_rows:
# #             ws.append(row)

# #         if data_rows:
# #             df_for_styling = pd.DataFrame(data_rows, columns=headers)
# #             apply_summary_sheet_styles(ws, df_for_styling)

# #     def style_row(self, sheet_name: str, row_number: int, is_header: bool = False):
# #         """指定した行にスタイルを適用するヘルパーメソッド"""
# #         if sheet_name not in self.workbook.sheetnames:
# #             return
# #         ws = self.workbook[sheet_name]

# #         if is_header:
# #             font = Font(bold=True, color=config.HEADER_FONT_COLOR)
# #             fill = PatternFill(start_color=config.HEADER_FILL_COLOR, end_color=config.HEADER_FILL_COLOR, fill_type="solid")
# #             for cell in ws[row_number]:
# #                 cell.font = font
# #                 cell.fill = fill

# #     def merge_header_cells(self, sheet_name: str, start_row: int, start_col: int, end_col: int):
# #         """指定された範囲のセルを結合する"""
# #         if sheet_name not in self.workbook.sheetnames:
# #             return
# #         ws = self.workbook[sheet_name]
# #         ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)

# #     def align_header_center(self, sheet_name: str, row_start: int, row_end: int):
# #         """指定された行範囲のセルを中央揃えにする"""
# #         if sheet_name not in self.workbook.sheetnames:
# #             return
# #         ws = self.workbook[sheet_name]
# #         center_alignment = Alignment(horizontal='center', vertical='center')
# #         for row in ws.iter_rows(min_row=row_start, max_row=row_end):
# #             for cell in row:
# #                 cell.alignment = center_alignment

# #     def apply_borders_to_all_cells(self, sheet_name: str):
# #         """
# #         指定されたシートのデータが存在する全てのセルに罫線を適用する。
# #         """
# #         if sheet_name not in self.workbook.sheetnames:
# #             return
# #         ws = self.workbook[sheet_name]

# #         thin_border_side = Side(style='thin', color='000000')
# #         border = Border(left=thin_border_side,
# #                         right=thin_border_side,
# #                         top=thin_border_side,
# #                         bottom=thin_border_side)

# #         for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
# #             for cell in row:
# #                 cell.border = border

# #     def _map_subject_columns(self, ws: Worksheet) -> dict:
# #         """成績一覧シートのヘッダー（科目）を解析し、(科目名, 種別) -> 列番号 の辞書を作成する"""
# #         mapping = {}
# #         fourth_row = ws[4]
# #         for cell_idx, cell_in_4th_row in enumerate(fourth_row):
# #             test_type = cell_in_4th_row.value
# #             if test_type in config.KEY_TEST_TYPES:
# #                 col_num_1based = cell_idx + 1
# #                 subject_name = self._get_merged_cell_value(ws, 1, col_num_1based)
# #                 if subject_name:
# #                     mapping[(subject_name, test_type)] = col_num_1based
# #         return mapping

# #     def _map_other_columns(self, ws: Worksheet) -> dict:
# #         """成績一覧シートのヘッダー（総点など）を解析し、項目名 -> 列番号 の辞書を作成する"""
# #         mapping = {}
# #         first_row = ws[1]
# #         for cell_idx, cell_in_1st_row in enumerate(first_row):
# #              if cell_in_1st_row.value in config.KEY_OTHER_COLS:
# #                  mapping[cell_in_1st_row.value] = cell_idx + 1
# #         return mapping

# #     def _get_merged_cell_value(self, ws: Worksheet, row: int, col: int) -> str:
# #         """
# #         指定されたセルが結合されている場合を考慮して、そのセルの表示上の値を取得する。
# #         """
# #         cell = ws.cell(row=row, column=col)
# #         if cell.value is not None:
# #             return cell.value

# #         current_search_col = col - 1
# #         while current_search_col >= 1:
# #             header_cell_to_check = ws.cell(row=row, column=current_search_col)
# #             if header_cell_to_check.value is not None:
# #                 return header_cell_to_check.value
# #             current_search_col -= 1
# #         return None





























